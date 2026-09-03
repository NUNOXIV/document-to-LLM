#!/usr/bin/env python3
"""Unit-Tests fuer die Teile der Pipeline, die ohne ML-Modelle pruefbar sind:
Qualitaetsgates, Seitenmarker-Export, Zielnamen, Pipeline-Optionen und die
Abweichungspruefung. Laeuft mit `python tests/test_units.py`.
"""
from __future__ import annotations

import sys
import tempfile
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

import extract  # noqa: E402
import gs_struktur as GS  # noqa: E402
import index as IDX  # noqa: E402
import tracker as TR  # noqa: E402
import publish  # noqa: E402
import versioncheck as VC  # noqa: E402
import verify as V  # noqa: E402

FIXTURE_PDF = Path(__file__).parent / "fixtures" / "Muster-Norm-Zweispaltig.pdf"
failures: list[str] = []


def check(name: str, cond: bool, detail: str = "") -> None:
    print(("  ok   " if cond else "  FAIL ") + name + (f" — {detail}" if detail and not cond else ""))
    if not cond:
        failures.append(name)


class FakeDoc:
    """Minimaler DoclingDocument-Ersatz fuer den seitenweisen Export."""

    def __init__(self, pages: dict[int, str]):
        self.pages = pages
        self.tables = []

    def export_to_markdown(self, page_no: int | None = None) -> str:
        if page_no is None:
            return "\n\n".join(self.pages[p] for p in sorted(self.pages))
        return self.pages[page_no]


class NoPageDoc(FakeDoc):
    def export_to_markdown(self, page_no: int | None = None) -> str:  # kein page_no
        if page_no is not None:
            raise TypeError("unexpected keyword argument 'page_no'")
        return "## Ohne Seiten\n\nText."


def test_page_markers() -> None:
    print("Seitenmarker")
    doc = FakeDoc({1: "## 4 Kontext\n\nText A.", 2: "## Anhang A\n\nText B."})
    md = extract.to_markdown(doc, with_page_markers=True)
    check("Marker Seite 1", "<!-- page: 1 -->" in md)
    check("Marker Seite 2", "<!-- page: 2 -->" in md)
    check("Reihenfolge", md.index("Text A.") < md.index("Text B."))
    check("Fallback ohne page_no", "Ohne Seiten" in extract.to_markdown(NoPageDoc({1: ""}), True))
    check("ohne Marker", "<!-- page:" not in extract.to_markdown(doc, False))


def test_zielname_kollision_gleiches_format() -> None:
    """Zwei Quellen mit gleichem Slug und gleichem Format duerfen sich nie
    ueberschreiben — so verschwanden 125 von 602 Extrakten, unbemerkt."""
    print("Zielnamen: Kollision im selben Format")
    with tempfile.TemporaryDirectory() as tmp:
        wurzel = Path(tmp)
        a = wurzel / "Checkliste-APP-1-1.xlsx"
        b = wurzel / "checklisten-2023" / "Checkliste_APP.1.1.xlsx"
        b.parent.mkdir()
        a.write_bytes(b"A")
        b.write_bytes(b"B")
        claimed: dict[str, Path] = {}
        na = extract.target_name(a, claimed)
        nb = extract.target_name(b, claimed)
        check("gleicher Prozess: verschiedene Ziele", na != nb, f"{na} / {nb}")
        check("Ausweichname nennt den Ordner, nicht das gleiche Format",
              nb == "checkliste-app-1-1-checklisten-2023", nb)
        check("gleicher Prozess: stabil", extract.target_name(b, claimed) == nb)
        # Zweiter Prozess: das Ziel von a liegt schon auf Platte, claimed ist leer.
        out = wurzel / "out"
        out.mkdir()
        (out / f"{na}.md").write_text(
            f'---\nsource_file: "{a.name}"\nsource_sha256: {extract.sha256_of(a)}\n---\n', encoding="utf-8")
        nb2 = extract.target_name(b, {}, out)
        check("anderer Prozess: fremdes Ziel nicht ueberschrieben", nb2 != na, nb2)
        check("anderer Prozess: gleicher Name wie im ersten", nb2 == nb, f"{nb2} / {nb}")
        check("eigenes Ziel wird wiedererkannt", extract.target_name(a, {}, out) == na)
        # Byte-identisches Duplikat (ZIP-Inhalt neben dem Original): kein zweites Ziel.
        c = wurzel / "kopie" / "Checkliste_APP-1-1.xlsx"
        c.parent.mkdir()
        c.write_bytes(b"A")
        check("Duplikat teilt das Ziel des Originals", extract.target_name(c, {}, out) == na)


def test_tabellenzeilen_ueberschreiben_keine_ueberschrift() -> None:
    """Anhang-A-Tabelle und Inhaltsverzeichnis duerfen die Kapitel nicht ueberschreiben.

    ISO 27001 und ISO 42001 fuehren Klausel 5.1 (Leadership) UND Control A.5.1
    (Policies). Die Tabellenzeile A.5.1 wurde auch unter "5.1" abgelegt und
    verdraengte die Ueberschrift: 32 Anforderungen trugen den Text einer
    anderen Nummer, und die Inhaltspruefung sah es nicht, weil sie dieselbe
    Zusammenfuehrung benutzte.
    """
    print("Tabellenzeilen gegen Ueberschriften")
    body = "\n".join([
        "## 5.1\tLeadership and commitment", "",
        "Top management shall demonstrate leadership.", "",
        "## 10\tImprovement", "",
        "## 10.1\tContinual improvement", "",
        "The organization shall continually improve the AI management system.", "",
        "| Clause | Title | Page |", "|---|---|---|",
        "| 10.1 | Continual improvement | 23 |", "",
        "| ID | Control | Text |", "|---|---|---|",
        "| A.5.1 | Policies for information security | Control: policies shall be defined. |",
        "| A.10 | Third-party and customer relationships | Third-party and customer relationships |",
        "| A.10.2 | Allocating responsibilities | The organization shall ensure that responsibilities are allocated. |",
        "| A.10.3 | Suppliers | The organization shall establish a supplier process. |",
        "", "## 10.2\tNonconformity", "", "When a nonconformity occurs, the organization shall react.", "",
        "## Annex A", "", "## Bibliography", "", "[1] ISO 31000",
    ])
    wanted = {"5.1": "Leadership and commitment", "A.5.1": "Policies for information security",
              "10.1": "Continual improvement", "10.2": "Nonconformity",
              "A.10": "Third-party and customer relationships",
              "A.10.2": "Allocating responsibilities", "A.10.3": "Suppliers"}
    out = publish.aufgeloeste_abschnitte(body, {}, wanted, "test")
    check("5.1 traegt den Kapiteltext", out["5.1"].text.startswith("Top management"), out["5.1"].text[:40])
    check("A.5.1 traegt den Control-Text", out["A.5.1"].text.startswith("Control:"), out["A.5.1"].text[:40])
    check("10.1 traegt den Kapiteltext, nicht die Inhaltsverzeichniszeile",
          out["10.1"].text.startswith("The organization shall continually"), out["10.1"].text[:40])
    check("A.10.2 traegt den Control-Text", "responsibilities are allocated" in out["A.10.2"].text)
    check("A.10 faellt nicht auf Kapitel 10 zurueck",
          "continually" not in out["A.10"].text and "Allocating" in out["A.10"].text, out["A.10"].text[:60])
    check("10.2 endet an der Strukturgrenze 'Annex A'",
          "react" in out["10.2"].text and "ISO 31000" not in out["10.2"].text, out["10.2"].text[:80])
    check("A.10 ist aus den Controls zusammengesetzt, nicht der wiederholte Titel",
          "responsibilities are allocated" in out["A.10"].text and "supplier process" in out["A.10"].text,
          out["A.10"].text[:80])


def test_vault_register_kennt_entfallene() -> None:
    """Withdrawn-Eintraege bleiben Sollwert (das Kompendium nennt sie noch), sind
    aber getrennt abfragbar, damit ihr Fehlen im Dokument kein Befund wird."""
    print("Vault-Register mit entfallenen Eintraegen")
    with tempfile.TemporaryDirectory() as tmp:
        ordner = Path(tmp) / "GRC" / "Frameworks" / "fw"
        ordner.mkdir(parents=True)
        (ordner / "fw X.1.md").write_text(
            "---\ntype: requirement\nid: X.1\nkind: requirement\n---\n# X.1 — Eins\n", encoding="utf-8")
        (ordner / "fw X.2.md").write_text(
            "---\ntype: requirement\nid: X.2\nkind: withdrawn\nstatus: withdrawn\n---\n# X.2 — ENTFALLEN\n",
            encoding="utf-8")
        ids = publish.vault_ids(Path(tmp), "fw")
        check("aktive ID im Register", "X.1" in ids and ids["X.1"] == "Eins", str(ids))
        check("entfallene ID bleibt Sollwert", "X.2" in ids, str(ids))
        check("entfallene ID getrennt abfragbar", publish.vault_withdrawn(Path(tmp), "fw") == {"X.2"})


def test_kreuzreferenz_grundschutz_druckfehler() -> None:
    """Druckfehler in der Quelle: die Anforderung wird ueber ihren Wortlaut gefunden,
    die Kennung im Extrakt bleibt, wie sie im Kompendium steht."""
    print("Kreuzreferenz Grundschutz")
    body = "\n".join([
        "## OPS.2.3.A21 Abschluss von ESCROW-Verträgen (H)", "", "Wird Software bezogen, SOLLTE ein ESCROW-Vertrag abgeschlossen werden.", "",
        "## OPS.2.3A22 Durchführung von gemeinsamen Notfall- und Krisenübungen (H) [Notfallbeauftragte]", "",
        "Gemeinsame Notfall- und Krisenübungen mit den Anbietenden von Outsourcing SOLLTEN durchgeführt und dokumentiert werden (siehe DER.4). Das Resultat SOLLTE genutzt werden.", "",
        "## OPS.2.3.A23 Einsatz von Verschlüsselungen (H)", "", "Sensible Daten SOLLTEN verschlüsselt werden.",
    ])
    wanted = {"OPS.2.3.A21": "Abschluss von ESCROW-Verträgen",
              "OPS.2.3.A22": "Durchführung von gemeinsamen Notfall- und Krisenübungen",
              "OPS.2.3.A23": "Einsatz von Verschlüsselungen"}
    out = publish.aufgeloeste_abschnitte(body, {}, wanted, "bsi-grundschutz")
    check("A22 ueber den Anker gefunden", "OPS.2.3.A22" in out, str(sorted(out)))
    check("'OPS.2.3A22' ist keine Ueberschrift der Gruppe OPS.2.3",
          "ops.2.3" not in publish.sections_from_headings(body)
          or "Gemeinsame" not in publish.sections_from_headings(body)["ops.2.3"].text)
    check("A22 traegt nur den eigenen Text",
          out["OPS.2.3.A22"].text.startswith("Gemeinsame") and "Sensible" not in out["OPS.2.3.A22"].text,
          out["OPS.2.3.A22"].text[:60])
    check("A23 unberuehrt", out["OPS.2.3.A23"].text.startswith("Sensible"))


def test_xberg_seitenmarken_normalisiert() -> None:
    """xberg klebt die Seitenmarke an das erste Wort der Seite; jeder Waechter
    sucht sie am Zeilenanfang."""
    print("xberg: Seitenmarken")
    roh = "<!-- page: 1 -->MUSTER-NORM\n\n# 4 Kontext\n\nText.\n\n<!-- page: 2 -->Anhang"
    md = extract.normalisiere_xberg_markdown(roh)
    zeilen = md.splitlines()
    check("Marke 1 allein auf der Zeile", "<!-- page: 1 -->" in zeilen, str(zeilen[:3]))
    check("Marke 2 allein auf der Zeile", "<!-- page: 2 -->" in zeilen, str(zeilen[-3:]))
    check("Text der Seite bleibt", "MUSTER-NORM" in zeilen and "Anhang" in zeilen)
    check("keine Dreifach-Leerzeilen", "\n\n\n" not in md)


def test_xberg_engine_am_fixture() -> None:
    """Zweite Engine am Test-PDF: gleicher Ausgabevertrag, gleiche Waechter.

    Laeuft nur, wenn xberg installiert ist. Geprueft wird der native Pfad
    (Textlayer, keine Modelle): Kopfzeile nennt die Engine, jede Seite hat ihre
    Marke, die Wortdeckung gegen den Textlayer ist vollstaendig.
    """
    print("xberg: Engine am Fixture")
    try:
        import xberg  # noqa: F401
    except ImportError:
        print("  (xberg nicht installiert — uebersprungen)")
        return
    if not FIXTURE_PDF.exists():
        print("  (Fixture fehlt — uebersprungen)")
        return
    with tempfile.TemporaryDirectory() as tmp:
        out = Path(tmp) / "out"
        res = extract.convert_file(
            extract._Runner(120), FIXTURE_PDF, out, ocr_mode="off", page_markers=True,
            write_json=True, force=True, claimed={}, do_verify=True,
            min_coverage=99.0, repair=True, engine="xberg",
        )
        check("Konverter xberg", res.converter == "xberg", res.converter)
        md = Path(res.output).read_text(encoding="utf-8")
        check("Kopfzeile nennt xberg", "converter: \"xberg " in md and "engine: xberg" in md)
        check("zwei Seitenmarken", md.count("<!-- page: 1 -->") == 1 and md.count("<!-- page: 2 -->") == 1)
        check("Ueberschriften erkannt", res.headings >= 5, str(res.headings))
        check("Wortdeckung vollstaendig", res.text_coverage is not None and res.text_coverage >= 99.0,
              str(res.text_coverage))
        check("JSON-Zwilling der Engine", res.json_output is not None and res.json_output.endswith(".xberg.json"),
              str(res.json_output))


def test_artikelgrenze_auch_ohne_ueberschrift() -> None:
    """Amtsblattsatz: Docling erkennt "Artikel 22" mal als Ueberschrift, mal als
    blosse Zeile. Ein Artikel endet trotzdem am naechsten Artikel -- sonst
    traegt Art.21 den Text von Art.22 und Art.23 mit (DSGVO: 9 Zeilen,
    DORA: 2, KI-VO: 1 -- vom Aufnahmetor des Auftraggebers gefunden, nicht
    von der Wortdeckung, denn doppelter Text hat volle Deckung)."""
    print("Artikelgrenze ohne Ueberschrift")
    body = "\n".join([
        "## Artikel 21", "", "## Widerspruchsrecht", "",
        "- (1) Die betroffene Person hat das Recht, Widerspruch einzulegen.", "",
        "Artikel 22", "", "Automatisierte Entscheidungen im Einzelfall", "",
        "(1) Die betroffene Person hat das Recht, nicht einer Entscheidung unterworfen zu werden.", "",
        "Artikel 23", "", "Beschränkungen", "",
        "(1) Durch Rechtsvorschriften koennen Pflichten beschraenkt werden.", "",
        "## KAPITEL IV", "",
        "## Artikel 61", "", "## Aenderung der Verordnung (EU) Nr. 909/2014", "",
        "Artikel 45 der Verordnung (EU) Nr. 909/2014 wird wie folgt geändert:", "",
        "1. Absatz 1 erhält folgende Fassung.",
    ])
    wanted = {"Art.21": "Widerspruchsrecht", "Art.22": "Automatisierte Entscheidungen",
              "Art.23": "Beschränkungen", "Art.45": "Vereinbarungen", "Art.61": "Aenderung"}
    out = publish.aufgeloeste_abschnitte(body, {}, wanted, "gdpr")
    check("Art.21 endet am blossen 'Artikel 22'",
          "Widerspruch" in out["Art.21"].text and "Automatisierte" not in out["Art.21"].text,
          out["Art.21"].text[:120])
    check("Art.22 hat eigenen Text und endet an 'Artikel 23'",
          "Art.22" in out and "unterworfen" in out["Art.22"].text and "beschraenkt" not in out["Art.22"].text,
          out.get("Art.22").text[:120] if "Art.22" in out else "fehlt")
    check("Art.23 gefunden", "Art.23" in out and "beschraenkt" in out["Art.23"].text)
    check("Art.45 ankert nicht an 'Artikel 45 der Verordnung ... wird geaendert'",
          "Art.45" not in out, out["Art.45"].text[:80] if "Art.45" in out else "")


def test_ueberhang_im_export() -> None:
    """Eine Zeile, die den vollen Text einer anderen Anforderung enthaelt, ist
    ein Ueberhang -- ausser die andere ist ihr Unterpunkt (9.2 aus 9.2.1)."""
    print("Ueberhang im Export")
    import inhalt
    b_text = "Die betroffene Person hat das Recht, nicht einer Entscheidung unterworfen zu werden, " * 3
    reqs = [
        {"id": "Art.21", "text": "Widerspruch. " * 10 + b_text},
        {"id": "Art.22", "text": b_text},
        {"id": "9.2", "text": "### 9.2.1 Allgemein\n\n" + "Die Organisation muss interne Audits durchfuehren. " * 5},
        {"id": "9.2.1", "text": "Die Organisation muss interne Audits durchfuehren. " * 5},
    ]
    gleich = "Sitzungen MUESSEN nach Inaktivitaet gesperrt werden. " * 6
    reqs += [{"id": "SYS.1.1.A31", "text": gleich}, {"id": "SYS.2.1.A33", "text": gleich},
             {"id": "SYS.1.1", "text": "### SYS.1.1.A31 Sperre\n\n" + gleich + "\n\nweiterer Text " * 20}]
    treffer = inhalt.ueberhaenge(reqs)
    check("Art.21 enthaelt Art.22", ("Art.21", "Art.22") in treffer, str(treffer))
    check("woertlich gleiche Anforderungen sind kein Ueberhang",
          not any(t[0] in ("SYS.1.1.A31", "SYS.2.1.A33") for t in treffer), str(treffer))
    check("Oberklausel mit dem Text ihres Unterpunkts trifft keine fremde Dublette",
          ("SYS.1.1", "SYS.2.1.A33") not in treffer, str(treffer))
    check("Oberklausel aus Unterpunkt ist kein Ueberhang", ("9.2", "9.2.1") not in treffer, str(treffer))


def test_quality_gates() -> None:
    print("Qualitaetsgates")
    try:
        extract.check_quality("---\na: 1\n---\n\n   \n", 2, True, False)
        check("leerer Extrakt bricht ab", False)
    except extract.ExtractionError:
        check("leerer Extrakt bricht ab", True)

    warn = extract.check_quality("# T\n\nx\n", 10, True, False)
    check("textarme Seiten gemeldet", any("Zeichen/Seite" in w for w in warn))
    check("OCR-Hinweis", any("--ocr on" in w for w in warn))

    flat = "Fliesstext ohne Gliederung. " * 40
    check("fehlende Ueberschriften gemeldet",
          any("Ueberschriften" in w for w in extract.check_quality(flat, 1, True, False)))

    broken = "# T\n\n| a | b |\n|---|---|\n| 1 | 2 | 3 |\n\n" + "Text. " * 60
    check("kaputte Tabelle gemeldet",
          any("Spaltenstruktur" in w for w in extract.check_quality(broken, 1, True, False)))

    good = "# T\n\n| a | b |\n|---|---|\n| 1 | 2 |\n\n" + "Text. " * 60
    check("intakte Tabelle ohne Warnung",
          not any("Spaltenstruktur" in w for w in extract.check_quality(good, 1, True, False)))

    check("Encoding-Fehler gemeldet",
          any("U+FFFD" in w for w in extract.check_quality("# T\n\nTe�xt " * 40, 1, True, False)))


def test_target_names() -> None:
    print("Zielnamen")
    claimed: dict[str, Path] = {}
    a = extract.target_name(Path("/in/ISO 27001.pdf"), claimed)
    b = extract.target_name(Path("/in/ISO 27001.docx"), claimed)
    check("Slug normalisiert", a == "iso-27001", a)
    check("keine Kollision", a != b and b == "iso-27001-docx", b)
    check("stabil bei Wiederholung", extract.target_name(Path("/in/ISO 27001.pdf"), claimed) == a)
    check("Umlaute", extract.slugify("Prüfung Änderung") == "pruefung-aenderung")


def test_pipeline_options() -> None:
    print("Docling-Pipelineoptionen")
    try:
        conv = extract.build_converter(ocr=True)
    except Exception as exc:  # ohne Modelle darf der Aufbau nicht scheitern
        check("Converter baubar", False, str(exc))
        return
    from docling.datamodel.base_models import InputFormat

    opts = conv.format_to_options[InputFormat.PDF].pipeline_options
    check("OCR gesetzt", opts.do_ocr is True)
    check("Tabellenstruktur an", opts.do_table_structure is True)
    check("Cell-Matching an", opts.table_structure_options.do_cell_matching is True)
    check("TableFormer ACCURATE", str(opts.table_structure_options.mode).upper().endswith("ACCURATE"),
          str(opts.table_structure_options.mode))
    check("OCR abschaltbar", extract.build_converter(ocr=False)
          .format_to_options[InputFormat.PDF].pipeline_options.do_ocr is False)


def test_verify() -> None:
    print("Abweichungspruefung")
    if not FIXTURE_PDF.exists():
        check("Fixture vorhanden", False, str(FIXTURE_PDF))
        return
    pages, boiler = V.pdf_pages(FIXTURE_PDF)
    check("Seiten gelesen", len(pages) == 2, str(len(pages)))
    check("Kopf-/Fusszeile erkannt", any("testzwecken" in t for t in boiler[1]))
    check("Kopfzeile nicht im Inhalt", not any("testzwecken" in t for t in pages[1]))

    with tempfile.TemporaryDirectory() as tmp:
        full = Path(tmp) / "full.md"
        full.write_text("---\nx: 1\n---\n" + "\n".join(" ".join(t) for t in pages.values()),
                        encoding="utf-8")
        r = V.verify(FIXTURE_PDF, full)
        check("vollstaendig = 100 %", r.coverage == 100.0, str(r.coverage))

        gap = Path(tmp) / "gap.md"
        gap.write_text(full.read_text(encoding="utf-8")
                       .replace("kryptographie", "").replace("schlüsselverwaltung", ""),
                       encoding="utf-8")
        rg = V.verify(FIXTURE_PDF, gap)
        check("Luecke erkannt", rg.coverage < 99.5, str(rg.coverage))
        check("fehlende Woerter benannt", "kryptographie" in rg.missing_sample)
        check("Seite lokalisiert", rg.per_page[2] < rg.per_page[1])

        onepage = Path(tmp) / "p2.md"
        onepage.write_text("---\nx: 1\n---\n" + " ".join(pages[2]), encoding="utf-8")
        r1 = V.verify(FIXTURE_PDF, onepage)
        check("fehlende Seite erkannt", r1.per_page[1] < 50, str(r1.per_page))

        joined = Path(tmp) / "join.md"
        joined.write_text(full.read_text(encoding="utf-8")
                          .replace("informationssicherheitsrichtlinie n",
                                   "informationssicherheitsrichtlinien"), encoding="utf-8")
        rj = V.verify(FIXTURE_PDF, joined)
        check("Umbruch-Zusammenfuehrung kein Fehlalarm", rj.coverage == 100.0, str(rj.coverage))


def test_repair() -> None:
    print("Nachtrag fuer nicht zugeordneten Text")
    if not FIXTURE_PDF.exists():
        check("Fixture vorhanden", False, str(FIXTURE_PDF))
        return
    pages, _ = V.pdf_pages(FIXTURE_PDF)
    with tempfile.TemporaryDirectory() as tmp:
        # Extrakt, dem der Tabellentext fehlt (wie ein verschluckter Zellrest)
        gap = Path(tmp) / "gap.md"
        gap.write_text("---\nx: 1\n---\n" + " ".join(pages[1]), encoding="utf-8")
        lines = V.unassigned_lines(FIXTURE_PDF, gap)
        check("fehlende Quellzeilen gefunden", bool(lines), str(len(lines)))
        check("Seitenzahl mitgeliefert", all(isinstance(p, int) for p, _ in lines))
        check("Zellinhalt enthalten",
              any("Schlüsselverwaltung" in text for _, text in lines))

        md = extract.appendix(lines)
        check("Nachtrag hat Ueberschrift", md.startswith("## Nachtrag"))
        check("Nachtrag mit Seitenmarker", "<!-- page: 2 -->" in md)
        check("Nachtrag als Zitat", "\n> " in md)

        # Nach dem Anhaengen muss die Deckung 100 % erreichen
        fixed = Path(tmp) / "fixed.md"
        fixed.write_text(gap.read_text(encoding="utf-8") + "\n\n" + md, encoding="utf-8")
        check("Deckung nach Nachtrag 100 %", V.verify(FIXTURE_PDF, fixed).coverage == 100.0,
              str(V.verify(FIXTURE_PDF, fixed).coverage))

        # Vollstaendiger Extrakt: kein Nachtrag noetig
        full = Path(tmp) / "full.md"
        full.write_text("---\nx: 1\n---\n" + "\n".join(" ".join(v) for v in pages.values()),
                        encoding="utf-8")
        check("kein Nachtrag bei vollstaendigem Extrakt", V.unassigned_lines(FIXTURE_PDF, full) == [])


def test_office_verify() -> None:
    print("Pruefung von Office-Formaten")
    import tempfile as tf
    from docx import Document as Docx
    from openpyxl import Workbook

    with tf.TemporaryDirectory() as tmp:
        xlsx = Path(tmp) / "controls.xlsx"
        wb = Workbook()
        ws = wb.active
        for row in [("Control", "Anforderung"),
                    ("A.8.24", "Regeln fuer Kryptographie und Schluesselverwaltung"),
                    ("A.8.25", "Regeln fuer sichere Entwicklung")]:
            ws.append(row)
        wb.save(xlsx)

        pages, _ = V.office_pages(xlsx)
        check("xlsx gelesen", bool(pages) and "schluesselverwaltung" in pages[1],
              str(list(pages)))

        full = Path(tmp) / "full.md"
        full.write_text("---\nx: 1\n---\n" + " ".join(pages[1]), encoding="utf-8")
        check("xlsx vollstaendig = 100 %", V.verify(xlsx, full).coverage == 100.0)

        gap = Path(tmp) / "gap.md"
        gap.write_text(full.read_text(encoding="utf-8").replace("schluesselverwaltung", ""),
                       encoding="utf-8")
        rg = V.verify(xlsx, gap)
        check("xlsx Luecke erkannt", rg.coverage < 100.0, str(rg.coverage))
        check("xlsx Nachtrag findet die Zelle",
              any("Schluesselverwaltung" in text for _, text in V.unassigned_lines(xlsx, gap)))

        docx = Path(tmp) / "policy.docx"
        d = Docx()
        d.add_paragraph("Die Organisation muss Kryptographie regeln.")
        d.save(docx)
        dpages, _ = V.office_pages(docx)
        check("docx gelesen", "kryptographie" in dpages[1], str(dpages))

    check("Tabellenkalkulation ohne Ueberschriften-Warnung",
          not any("Ueberschriften" in w for w in
                  extract.check_quality("| a | b |\n|---|---|\n| 1 | 2 |\n" + "Text " * 60,
                                        1, False, False, ".xlsx")))


def test_broken_ooxml_styles() -> None:
    """Web-Exporte schreiben `<fill />` ohne patternFill — openpyxl bricht ab.

    Geprueft wird, dass die Container-Normalisierung die Datei lesbar macht
    und dabei kein Zellinhalt verloren geht.
    """
    import re
    import tempfile as tf
    import zipfile
    from openpyxl import Workbook, load_workbook

    with tf.TemporaryDirectory() as tmp:
        tmpdir = Path(tmp)
        good = tmpdir / "good.xlsx"
        wb = Workbook()
        ws = wb.active
        ws.append(("Control", "Anforderung"))
        ws.append(("A.8.24", "Regeln fuer Kryptographie"))
        wb.save(good)

        # styles.xml gezielt beschaedigen: patternFill aus den fills entfernen
        broken = tmpdir / "broken.xlsx"
        with zipfile.ZipFile(good) as zin, \
                zipfile.ZipFile(broken, "w", zipfile.ZIP_DEFLATED) as zout:
            for item in zin.infolist():
                data = zin.read(item.filename)
                if item.filename.endswith("styles.xml"):
                    text = data.decode("utf-8")
                    text = re.sub(r"<fill>.*?</fill>", "<fill />", text, flags=re.S)
                    data = text.encode("utf-8")
                zout.writestr(item, data)

        try:
            load_workbook(broken, read_only=True, data_only=True)
            check("defektes Stylesheet bricht openpyxl ab", False, "kein Fehler")
        except TypeError:
            check("defektes Stylesheet bricht openpyxl ab", True)

        outdir = tmpdir / "fix"
        outdir.mkdir()
        repaired = V.normalize_ooxml_styles(broken, outdir)
        check("Reparatur liefert eine Kopie", repaired is not None and repaired.exists())

        readable, _tmp = V.readable_office_source(broken)
        pages, _ = V.office_pages(broken)
        check("repariertes xlsx wird gelesen",
              bool(pages) and "kryptographie" in pages[1], str(pages))
        check("Original bleibt unveraendert", readable != broken)

        # An intakten Dateien darf nichts repariert werden.
        check("intakte Datei wird nicht angefasst",
              V.normalize_ooxml_styles(good, outdir) is None)


def test_text_passthrough() -> None:
    """Formate ohne Docling-Reader werden zeichengetreu uebernommen."""
    print("\ntest_text_passthrough")
    with tempfile.TemporaryDirectory() as td:
        d = Path(td)
        # Inhalt mit Backticks, damit der Zaun wachsen muss, und mit CRLF.
        content = "a: 1\r\n```\r\nnoch ein: wert\r\n"
        src = d / "kriterien.yml"
        src.write_bytes(content.encode("utf-8"))

        body, lines = extract.passthrough_body(src)
        check("Zeilen gezaehlt", lines == 3, f"{lines}")
        check("Zaun waechst ueber Quell-Backticks", "````yaml" in body, body[:80])

        fence = "`" * 4
        inner = body.split(fence + "yaml\n", 1)[1].rsplit("\n" + fence, 1)[0]
        check("Inhalt zeichengetreu", inner == "a: 1\n```\nnoch ein: wert", repr(inner))

        # .yml gilt als unterstuetzt, laeuft aber nicht ueber Docling.
        check("Suffix unterstuetzt", ".yml" in extract.SUPPORTED_SUFFIXES)
        check("nicht im Docling-Pfad", ".yml" not in extract.DOCLING_SUFFIXES)
        # .mm (FreeMind-Mindmap) ist XML und laeuft ueber denselben Pfad.
        check("mm unterstuetzt", ".mm" in extract.SUPPORTED_SUFFIXES)
        check("mm als XML eingezaeunt", extract.TEXT_FENCE_LANG[".mm"] == "xml")

        out = d / "out"
        res = extract.convert_file(
            None, src, out, ocr_mode="auto", page_markers=True,
            write_json=False, force=True, claimed={}, do_verify=True,
            min_coverage=99.0, repair=True,
        )
        check("Konverter markiert", res.converter == "passthrough", res.converter)
        check("Deckung 100 %", res.text_coverage == 100.0, str(res.text_coverage))
        check("Warnung gesetzt", any("keinen Reader" in w for w in res.warnings))
        md = Path(res.output).read_text(encoding="utf-8")
        check("Front-Matter nennt Passthrough", "ACSOS Passthrough" in md)
        check("Quelltext im Extrakt", "noch ein: wert" in md)



def test_yaml_catalogue() -> None:
    """Maschinenlesbarer Katalog (BSI C5 als YAML) -> Anforderungsabschnitte."""
    print("\ntest_yaml_catalogue")

    flach = """<!-- irgendein Vorspann -->
```yaml
-
  id: 'GC-01'
  name: 'Angaben zum Recht'
  condition: 'Der Anbieter macht Angaben zu:\n\n1. dem anwendbaren Recht.'
  hint: 'Vgl. Abschnitt 1.2.'
```
"""
    got = publish.sections_from_yaml(flach, {"source_file": "GC.yml"})
    check("flache Form gefunden", "gc-01" in got, str(sorted(got)))
    sec = got.get("gc-01")
    check("Titel aus name", bool(sec) and sec.title == "Angaben zum Recht")
    check("Normtext woertlich", bool(sec) and "dem anwendbaren Recht." in sec.text)
    check("Hinweis getrennt", bool(sec) and "[!info] Hinweis der Quelle" in sec.text)
    check("Schluesselpfad statt Seite",
          bool(sec) and sec.page == 0 and sec.locator == "GC-01 in GC.yml")

    verschachtelt = """```yaml
-
  identifier: '01'
  name: 'Rahmenwerk'
  basic:
    -
      identifier: '01B'
      criterion: 'Ein Rahmenwerk ist dokumentiert.'
  additional_complement:
    -
      identifier: '01AC'
      criterion: 'Zusaetzlich werden Informationen beruecksichtigt.'
  information:
    -
      information_text: 'Assets sind Objekte im Verantwortungsbereich.'
```
"""
    got = publish.sections_from_yaml(verschachtelt, {"source_file": "AM.yml"})
    check("Unterkriterium Basic", "am-01.01b" in got, str(sorted(got)))
    check("Unterkriterium Complement", "am-01.01ac" in got)
    check("Kriterienbereich aus information", "am-01" in got)
    check("Gruppe aus Dateiname",
          got.get("am-01.01b") is not None
          and got["am-01.01b"].text.startswith("Ein Rahmenwerk"))
    check("Belegstelle nennt Zweig",
          got.get("am-01.01b") is not None
          and got["am-01.01b"].locator == "basic/01B in AM.yml")

    # Kein Katalog: Passthrough einer Konfigurationsdatei darf nichts liefern.
    check("Nicht-Katalog ergibt nichts",
          publish.sections_from_yaml("```yaml\nversion: '1.1.0'\n```\n",
                                     {"source_file": "Version-und-Lizenz.yml"}) == {})
    check("ohne YAML-Block nichts",
          publish.sections_from_yaml("# Ohne Block\n", {"source_file": "AM.yml"}) == {})

    # Nur ein Katalog darf das Fehlen einer ID belegen.
    check("Katalog nennt seine Gruppe",
          publish.yaml_catalogue_group(verschachtelt, {"source_file": "AM.yml"}) == "AM")
    check("PDF-Extrakt ist kein Katalog",
          publish.yaml_catalogue_group("## 4.1 Kontext\n\nText.\n",
                                       {"source_file": "norm.pdf"}) is None)
    note = publish.withdrawn_note("bsi-c5", "AM-01.02AC", {"source_file": "AM.yml"})
    check("Entfallen-Notiz kennzeichnet Status", "status: entfallen" in note)
    check("Entfallen-Notiz erfindet keinen Text", "> [!quote]" not in note)
    check("Entfallen-Notiz warnt vor Zitat", "Nicht als geltende Anforderung zitieren" in note)

    # Ueberholte Fassungen: Wortlaut bleibt, Geltung nicht.
    with tempfile.TemporaryDirectory() as tmp:
        d = Path(tmp)
        src = d / "alt.md"
        src.write_text("# Alt\n\nAlter Wortlaut.\n", encoding="utf-8")
        vault = d / "vault"
        (vault / "GRC" / "Handbuch").mkdir(parents=True)
        m = {"source_file": "Alt.pdf", "source_sha256": "ab" * 32, "pages": "3"}
        note, full = publish.document_notes(src, vault, m, "Alter Wortlaut.",
                                            "Alte Fassung", "BSI", "Leitfaden",
                                            dry_run=False, superseded_by="neu-2026")
        txt = note.read_text(encoding="utf-8")
        check("Historie: Status gesetzt", "status: historisch" in txt)
        check("Historie: Nachfolger genannt", "superseded_by: neu-2026" in txt)
        check("Historie: warnt vor Zitat", "nicht als geltend zitieren" in txt)
        check("Historie: eigener Tag", "grc/historisch" in txt)
        check("Historie: Wortlaut bleibt", "Alter Wortlaut." in full.read_text(encoding="utf-8"))

        note2, _ = publish.document_notes(src, vault, m, "Alter Wortlaut.",
                                          "Normale Fassung", "BSI", "Leitfaden",
                                          dry_run=False)
        t2 = note2.read_text(encoding="utf-8")
        check("ohne Nachfolger keine Warnung", "status: historisch" not in t2)



def test_gs_struktur() -> None:
    """Mindmap-Hierarchie und Plakat-Legende zu einer Gliederung fuegen."""
    print("\ntest_gs_struktur")
    import xml.etree.ElementTree as ET

    mm = ET.fromstring(
        '<map><node TEXT="Wurzel">'
        '<node TEXT="APP (Anwendungen)">'
        '<node TEXT="APP.1.1 Office"><icon BUILTIN="full-2"/></node>'
        '<node TEXT="APP.1.9 Ohne Symbol"/>'
        "</node>"
        '<node><richcontent TYPE="NODE"><html><body><p>SYS.1.3 Linux</p>'
        "</body></html></richcontent><icon BUILTIN=\"full-1\"/></node>"
        "</node></map>")
    wurzel = mm.find("node")

    zeilen, ohne = [], []
    GS.gliederung(wurzel, 0, zeilen, ohne)
    check("Hierarchie eingerueckt", zeilen[1].startswith("- APP (") and
          zeilen[2].startswith("  - APP.1.1"), str(zeilen[:3]))
    check("Rang aus Symbol", "· **R2**" in zeilen[2], zeilen[2])
    check("richcontent gelesen", any("SYS.1.3 Linux" in z for z in zeilen), str(zeilen))
    check("Rang R1 erkannt", any("SYS.1.3 Linux · **R1**" in z for z in zeilen))
    # Ein rangloses Blatt ist auffaellig, ein Gruppenknoten nicht.
    check("rangloses Blatt gemeldet", ohne == ["APP.1.9 Ohne Symbol"], str(ohne))

    with tempfile.TemporaryDirectory() as tmp:
        plakat = Path(tmp) / "plakat.md"
        plakat.write_text(
            "> IT-Grundschutz-Kompendium 2023\n"
            "> : Zuerst umsetzen.\n> : Danach umsetzen.\n> : Zuletzt umsetzen.\n"
            "> Farben: neuer Baustein\n> Stand: Februar 2023\n", encoding="utf-8")
        erkl, fuss = GS.legende(plakat)
        check("drei Rangstufen erklaert", erkl == {1: "Zuerst umsetzen.",
                                                  2: "Danach umsetzen.",
                                                  3: "Zuletzt umsetzen."}, str(erkl))
        check("Fusszeilen erkannt", fuss == ["Farben: neuer Baustein",
                                             "Stand: Februar 2023"], str(fuss))



def test_versioncheck_historie() -> None:
    """Bewusst gefuehrte Altfassungen sind kein Mangel."""
    print("\ntest_versioncheck_historie")
    hist = VC.historisch()
    check("Historienregistry gelesen", isinstance(hist, dict) and hist, str(hist)[:80])
    check("OWASP 4.0 als historisch gefuehrt",
          any("owasp" in s for s in hist), str(list(hist)[:3]))
    check("Nachfolger benannt", all(v for v in hist.values()), str(hist))

    # Der Bericht darf eine Altfassung nicht als 'veraltet' zaehlen.
    f_hist = VC.Finding("alt", "Alt", "4.0", "4.2", "u", "historisch", "", "")
    f_alt = VC.Finding("x", "X", "1.0", "2.0", "u", "veraltet", "", "")
    text = VC.render([f_hist, f_alt])
    check("nur echte Veraltung gezaehlt", "veraltet: 1" in text, text[:200])
    check("historisch im Bericht benannt", "historisch" in text)



def test_fts5_query() -> None:
    """Alltagsschreibweisen duerfen nicht als FTS5-Syntax gelesen werden."""
    print("\ntest_fts5_query")
    q = IDX.fts5_query
    # Der Bindestrich im Wort war der Fehler: FTS5 las ihn als Spaltenfilter.
    check("Bindestrich gequotet", q("Mehrfaktor-Authentisierung Fernzugriff")
          == '"Mehrfaktor-Authentisierung" "Fernzugriff"', q("Mehrfaktor-Authentisierung Fernzugriff"))
    check("Punkte in IDs gequotet", q("OPS.1.1.3") == q("OPS.1.1.3"))
    check("Schraegstrich gequotet", '"IT/OT"' in q("IT/OT Netz") or q("IT/OT Netz") == "IT/OT Netz")
    # Ohne Sonderzeichen bleibt die Eingabe unangetastet.
    check("harmlose Eingabe unveraendert", q("Backup Konzept") == "Backup Konzept")
    # Bewusste FTS5-Syntax wird nicht entschaerft.
    check("Phrase bleibt", q('"genau so"') == '"genau so"')
    check("OR bleibt", q("Backup OR Sicherung") == "Backup OR Sicherung")
    check("NEAR bleibt", q("NEAR(Backup Test)") == "NEAR(Backup Test)")



def test_korpus_json() -> None:
    """Bestandsregister: ein verarbeitendes System darf nicht raten muessen."""
    print("\ntest_korpus_json")
    import json as J
    with tempfile.TemporaryDirectory() as tmp:
        out = Path(tmp)
        (out / "a.docling.json").write_text("{}", encoding="utf-8")
        (out / "b.passthrough.json").write_text("{}", encoding="utf-8")
        docs = [TR.Doc(slug="a", source="a.pdf", coverage=100.0, pages=3),
                TR.Doc(slug="b", source="b.yml", coverage=100.0),
                TR.Doc(slug="c", source="c.pdf", coverage=99.0),
                TR.Doc(slug="scan", source="scan.pdf", coverage=None, ocr=True)]
        d = J.loads(TR.korpus_json(docs, out))
    check("alle Dokumente gelistet", d["documents_total"] == 4, str(d["documents_total"]))
    arten = {e["slug"]: e["struktur_art"] for e in d["documents"]}
    check("Docling erkannt", arten["a"] == "docling", str(arten))
    check("Passthrough erkannt", arten["b"] == "passthrough", str(arten))
    # Ohne Struktur-JSON darf nicht stillschweigend fehlen: es wird benannt.
    check("fehlende Struktur benannt", d["ohne_struktur_json"] == ["c", "scan"],
          str(d["ohne_struktur_json"]))
    check("kein Pfad erfunden", arten["c"] is None
          and d["documents"][2]["struktur_json"] is None)
    check("Befund mitgefuehrt", d["documents"][0]["befund"] == "vollstaendig",
          d["documents"][0]["befund"])
    # Ein Scan darf nicht wie ein woertlicher Extrakt aussehen: ein leeres
    # Deckungsfeld allein liesse sich als "unbekannt" lesen. Die Flags sagen
    # ausdruecklich, dass der Text geraten und nicht gelesen wurde.
    nach_slug = {e["slug"]: e for e in d["documents"]}
    check("Scan als nicht woertlich markiert",
          nach_slug["scan"]["woertlich"] is False and nach_slug["scan"]["ocr"] is True,
          str(nach_slug["scan"]))
    check("Extrakt mit Textlayer bleibt woertlich",
          nach_slug["a"]["woertlich"] is True and nach_slug["a"]["ocr"] is False,
          str(nach_slug["a"]))



def test_export_json() -> None:
    """Anforderungs-Export in der Form, die ein rechnendes System erwartet."""
    print("\ntest_export_json")
    import json as J
    g = publish.gruppe
    check("A.5.1 -> A.5", g("A.5.1") == "A.5", g("A.5.1"))
    check("APP.1.1.A1 -> APP.1.1", g("APP.1.1.A1") == "APP.1.1", g("APP.1.1.A1"))
    check("AM-01.01B -> AM-01", g("AM-01.01B") == "AM-01", g("AM-01.01B"))
    check("GC-01 -> GC", g("GC-01") == "GC", g("GC-01"))
    check("ohne Gliederung leer", g("PO") == "", g("PO"))

    treffer = {"A.5.1": publish.Section("A.5.1", "Policies", "Wortlaut.", 12)}
    d = J.loads(publish.export_json("iso27001-2022", "unbekannt",
                                    {"source_file": "n.pdf", "source_sha256": "ab"},
                                    treffer, ["A.5.2"]))
    check("Pflichtfelder vorhanden",
          set(d) >= {"frameworkId", "edition", "sourceFile", "sourceSha256",
                     "requirements", "missing"}, str(sorted(d)))
    r = d["requirements"][0]
    check("Anforderung vollstaendig",
          r == {"id": "A.5.1", "title": "Policies", "text": "Wortlaut.", "group": "A.5"}, str(r))
    # Nicht aufgeloeste IDs duerfen nicht als Anforderung ohne Wortlaut erscheinen.
    check("Luecke getrennt gefuehrt", d["missing"] == ["A.5.2"]
          and all(x["text"] for x in d["requirements"]), str(d["missing"]))
    check("keine Ausgabe erfunden", d["edition"] is None, str(d["edition"]))


def main() -> int:
    """Laeuft ohne pytest — findet die Tests aber selbst.

    Vorher stand hier eine handgepflegte Liste. Jeder danach geschriebene Test
    fehlte darin und lief im CI-Skriptpfad nie mit: fuenf Waechtertests waren
    vorhanden und wurden nicht ausgefuehrt. Eine Liste, die man vergessen kann,
    ist keine Pruefung. Tests mit Fixtures (tmp_path) braucht pytest; sie werden
    hier ausdruecklich als nicht ausgefuehrt benannt statt stillschweigend
    uebergangen.
    """
    import inspect

    hier = sys.modules[__name__]
    alle = [(n, f) for n, f in vars(hier).items()
            if n.startswith("test_") and inspect.isfunction(f)]
    alle.sort(key=lambda nf: nf[1].__code__.co_firstlineno)
    ohne_fixture = [(n, f) for n, f in alle if not inspect.signature(f).parameters]
    mit_fixture = [n for n, f in alle if inspect.signature(f).parameters]
    for _, fn in ohne_fixture:
        fn()
    print()
    print(f"{len(ohne_fixture)} Test(s) ohne Fixture ausgefuehrt.")
    if mit_fixture:
        print(f"{len(mit_fixture)} Test(s) brauchen pytest und liefen hier NICHT: "
              + ", ".join(mit_fixture))
    if failures:
        print(f"{len(failures)} Test(s) fehlgeschlagen: {', '.join(failures)}")
        return 1
    print("Alle Unit-Tests bestanden.")
    return 0


def test_document_note_ohne_pruefbare_deckung(tmp_path: Path) -> None:
    """Ein Scan ohne Textlayer darf nicht wie ein geprueftes Dokument aussehen.

    Frueher stand in der Notiz "Wortdeckung — %": das liest sich wie ein
    Formatierungsfehler, nicht wie eine Aussage. Genau hier muss aber stehen,
    dass der Text aus der Zeichenerkennung stammt und niemand ihn gegen ein
    Original gehalten hat — sonst wandert OCR-Text als belegter Wortlaut in
    ein Audit.
    """
    import publish

    vault = tmp_path / "vault"
    (vault / "GRC" / "Handbuch").mkdir(parents=True)
    (vault / publish.LICENSED_DIR / "dokumente").mkdir(parents=True)
    md = tmp_path / "scan-ohne-textlayer.md"
    md.write_text("# Titel\n\nText aus OCR.\n", encoding="utf-8")

    meta = {"source_file": "Scan.pdf", "source_sha256": "abc", "pages": "14",
            "text_coverage_percent": "", "converter": "IBM Docling 2.123.0"}
    note, _ = publish.document_notes(md, vault, meta, "Text aus OCR.",
                                     "Titel", "Autor", "Foliensatz", False)
    t = note.read_text(encoding="utf-8")
    assert "Wortdeckung —" not in t, "kein Strich als Scheinwert"
    assert "Maschinell gelesen" in t, "Warnhinweis fehlt"
    assert "laesst sich nicht berechnen" in t
    assert "text_coverage_percent: null" in t
    assert "deckung_pruefbar: false" in t
    assert "erzeugt, nicht" in t and "extrahiert" in t

    meta["text_coverage_percent"] = "100.0"
    note2, _ = publish.document_notes(md, vault, meta, "Text.",
                                      "Titel2", "Autor", "Fachbuch", False)
    t2 = note2.read_text(encoding="utf-8")
    assert "Wortdeckung 100.0 %" in t2
    assert "deckung_pruefbar" not in t2


def test_deckung_braucht_tragfaehige_grundlage(tmp_path: Path) -> None:
    """Eine Deckungszahl gegen zwei Woerter ist kein Befund, sondern Zufall.

    Beobachtet an einem 291-Seiten-Scan: dessen Textlayer enthielt zwei
    Streuzeichen aus einer Kopfzeile. Der Vergleich fand beide im Extrakt
    wieder und meldete 100,0 % — fuer ein Dokument, dessen 195773 Woerter
    saemtlich aus der Zeichenerkennung stammten. Im Tracker haette es als
    "vollstaendig" gestanden, im Korpusregister als woertlich.

    Die alte Schutzklausel griff nur bei einer voellig leeren Grundlage.
    """
    import verify as V

    class Fake:
        def __init__(self, ref: int, ext: int) -> None:
            self.ref, self.ext = ref, ext

    def lauf(ref_woerter: int, ext_woerter: int) -> V.VerifyResult:
        md = tmp_path / f"e{ref_woerter}-{ext_woerter}.md"
        md.write_text(" ".join(f"wort{i}" for i in range(ext_woerter)), encoding="utf-8")
        quelle = tmp_path / f"q{ref_woerter}-{ext_woerter}.pdf"
        quelle.write_bytes(b"%PDF-1.4\n")
        seiten = {1: [f"wort{i}" for i in range(ref_woerter)]}
        echt = V.source_pages
        V.source_pages = lambda p: (seiten, {})
        try:
            return V.verify(quelle, md)
        finally:
            V.source_pages = echt

    duenn = lauf(2, 5000)
    assert duenn.coverage is None, "keine Zahl ohne tragfaehige Grundlage"
    assert "zu duenn" in (duenn.note or ""), duenn.note

    # Gegenprobe: ein kleines Dokument ist nicht verdaechtig, nur weil es klein
    # ist. Sonst faellt jede winzige Tabelle faelschlich in denselben Zweig.
    klein = lauf(30, 30)
    assert klein.coverage == 100.0, str(klein.coverage)
    assert not klein.note

    gesund = lauf(4800, 5000)
    assert gesund.coverage is not None and gesund.coverage > 90.0

    # Office-Formate bleiben aussen vor. Ein XLSX mit verbundenen Zellen
    # blaeht den Extrakt durch reine Wiederholung auf -- A3_Modellierung
    # kommt auf 4995 Quellwoerter gegen 114056 Extraktwoerter, ein
    # Verhaeltnis von 4,4 %, ohne dass ein einziges Wort ungeprueft waere.
    # Der Schutz gilt dem OCR-Fall, und OCR gibt es nur bei PDFs.
    md = tmp_path / "tabelle.md"
    md.write_text(" ".join(f"wort{i%20}" for i in range(5000)), encoding="utf-8")
    xlsx = tmp_path / "tabelle.xlsx"
    xlsx.write_bytes(b"PK\x03\x04")
    echt = V.source_pages
    V.source_pages = lambda p: ({1: [f"wort{i%20}" for i in range(200)]}, {})
    try:
        office = V.verify(xlsx, md)
    finally:
        V.source_pages = echt
    assert office.coverage == 100.0, str(office.coverage)
    assert not office.note, office.note


def test_tracker_befund_trennt_duenn_von_fehlend() -> None:
    """Ein zu duenner Textlayer ist nicht dasselbe wie gar keiner.

    Beide fuehren zu 'keine Deckungszahl', aber nur der zweite Fall ist
    offensichtlich. Der erste gab sich vorher als volle Deckung aus und
    gehoert deshalb im Protokoll beim Namen genannt, statt unter dem
    harmloseren Befund mitzulaufen.
    """
    import tracker as TR

    duenn = TR.Doc(slug="scan", coverage=None,
                   warnings=["Textlayer zu duenn fuer einen Wortvergleich: 2 "
                             "Referenzwoerter gegen 195773 Woerter im Extrakt."])
    leer = TR.Doc(slug="bild", coverage=None,
                  warnings=["Kein Textlayer im PDF (gescannt)"])
    voll = TR.Doc(slug="ok", coverage=100.0)

    assert duenn.verdict == "Textlayer zu duenn", duenn.verdict
    assert leer.verdict == "kein Textlayer", leer.verdict
    assert voll.verdict == "vollstaendig", voll.verdict

    # Dritter Fall: Markdown-Quelle. Kein zweiter Leser, also keine
    # Gegenprobe -- aber ein Textlayer fehlt hier nicht, es gibt schlicht
    # nur Text. "kein Textlayer" waere eine Falschaussage.
    md_quelle = TR.Doc(slug="md", coverage=None, warnings=[])
    assert md_quelle.verdict == "nicht gegengeprueft (Format)", md_quelle.verdict


def test_tracker_vault_luecken() -> None:
    """Fehlende Anforderungen im Vault-Geruest gehoeren ins Protokoll.

    Sie sind kein Extraktionsfehler: der Wortlaut liegt vor, nur das Raster
    kennt die ID nicht. Genau deshalb faellt so etwas sonst niemandem auf —
    eine Abdeckungsanalyse gegen ein unvollstaendiges Raster meldet keine
    Luecke, sie hat die Anforderung nie gesehen.
    """
    import tracker as TR

    g = TR.vault_gaps()
    assert g.get("offen_nach_korrektur"), "Registry ist leer"
    for e in g["offen_nach_korrektur"]:
        assert e.get("id") and e.get("befund"), str(e)
    assert g.get("hinweis") and g.get("ursache") and g.get("lehre")
    # Der Eintrag haelt eine Korrektur fest: der frueher gemeldete Mangel am
    # Vault existierte nicht, verglichen wurde gegen die falsche Menge. Das
    # gehoert benannt, nicht stillschweigend ersetzt.
    assert "KORREKTUR" in g["hinweis"]


def test_export_laengen_plausibel(tmp_path: Path) -> None:
    """Ein befuelltes Feld ist noch kein richtiges Feld.

    Der Grundschutz-Export trug im Median 54019 Zeichen je Anforderung statt
    der ueblichen paar hundert: die Ueberschriftenerkennung kannte keine
    Kennungen mit Buchstabenpraefix (APP.1.1.A1), fiel auf den Textanker
    zurueck, und der findet kein Ende — jede Anforderung schleppte den Rest
    des Dokuments mit. Die damalige Pruefung sah nur nach, ob ein Text da ist,
    ob IDs fehlen und ob welche doppelt sind. Alles gruen, alles falsch.

    Deshalb prueft dieser Test die Verteilung, nicht die Anwesenheit.
    """
    import publish

    body = "\n".join([
        "## APP.1.1.A1 Erste Anforderung (B)", "",
        "Die Institution MUSS das eine tun.", "",
        "## APP.1.1.A2 Zweite Anforderung (S)", "",
        "Die Institution SOLLTE das andere tun.", "",
        "## SYS.2.2.3.A7 Dritte Anforderung (H)", "",
        "Die Institution KANN das dritte tun.", "",
    ])
    s = publish.sections_from_headings(body)
    for ident in ("app.1.1.a1", "app.1.1.a2", "sys.2.2.3.a7"):
        assert ident in s, f"{ident} nicht erkannt — Buchstabenpraefix faellt durch"

    # INF, IND und ISMS beginnen mit Buchstaben, die auch roemische Ziffern
    # sind. Stand der roemische Zweig im Regex vorn, matchte er das blosse "I"
    # und lieferte die Kennung "I" statt "INF.1.A1" -- 49 Anforderungen
    # verloren dadurch ihre Abschnittsgrenze, waehrend der Median gesund
    # aussah. Ein Test nur mit APP und SYS haette das durchgelassen.
    roem = "\n".join([
        "## INF.1.A1 Planung der Gebaeudeabsicherung (B)", "",
        "Die Institution MUSS planen.", "",
        "## IND.2.1.A3 Nutzung sicherer Protokolle (S)", "",
        "Die Institution SOLLTE sichere Protokolle nutzen.", "",
        "## ISMS.1.A6 Integration in Ablaeufe (B)", "",
        "Die Institution MUSS integrieren.", "",
        "## X Nachfolgende Ueberschrift", "", "Ende.", "",
    ])
    sr = publish.sections_from_headings(roem)
    for ident in ("inf.1.a1", "ind.2.1.a3", "isms.1.a6"):
        assert ident in sr, f"{ident} nicht erkannt — roemischer Zweig greift zu frueh"
        assert len(sr[ident].text) < 120, \
            f"{ident} laeuft ueber seine Grenze hinaus: {len(sr[ident].text)}"
    assert "I" not in sr, "blosses 'I' als Kennung erkannt"

    # Die entscheidende Zusicherung: ein Abschnitt endet an der naechsten
    # Ueberschrift. Ohne sie sieht der Export vollstaendig aus und ist es nicht.
    assert "das andere" not in s["app.1.1.a1"].text, \
        "Abschnitt laeuft in die naechste Anforderung hinein"
    assert "das dritte" not in s["app.1.1.a2"].text
    laengen = [len(s[i].text) for i in ("app.1.1.a1", "app.1.1.a2", "sys.2.2.3.a7")]
    assert max(laengen) < 200, f"unplausibel lang: {laengen}"


def test_zellversatz_repariert() -> None:
    """Text, der vor der Zellmarke steht, gehoert der vorigen Anforderung.

    Beobachtet in ISO/IEC 27001 Anhang A: A.5.16 trug nur "Control", waehrend
    ihr Satz am Anfang der Zelle von A.5.17 stand. Neun von 94 Zeilen waren so
    verschoben. Laengen, Kennungen und Feldbelegung bleiben dabei voellig
    unauffaellig — auffallen kann es nur, wer den Wortlaut gegen die Nummer
    haelt. In einem Compliance-Bestand ist das der teuerste Fehler: eine
    Anforderung, die etwas anderes sagt, als ihre Nummer verspricht.
    """
    import publish

    body = "\n".join([
        "| 5.15 | Access control | Control Rules to control access shall be established. |",
        "| 5.16 | Identity management | Control |",
        "| 5.17 | Authentication information | The full life cycle of identities shall be "
        "managed. Control Allocation of authentication information shall be controlled. |",
        "| 5.18 | Access rights | Control Access rights shall be provisioned. |",
        "| 5.19 | Supplier relationships | Control Processes shall be defined. |",
        "| 5.20 | Supplier agreements | Control Requirements shall be agreed. |",
    ])
    s = publish.sections_from_tables(body)
    assert "full life cycle of identities" in s["5.16"].text, s["5.16"].text
    assert "full life cycle" not in s["5.17"].text, s["5.17"].text
    assert "Allocation of authentication" in s["5.17"].text
    # Unverschobene Zeilen bleiben unangetastet.
    assert s["5.18"].text.startswith("Control Access rights")

    # Ohne erkennbare Marke wird nichts verschoben: lieber unrepariert als
    # falsch repariert.
    ohne = "\n".join([
        "| 1.1 | Alpha | Erster Text ohne gemeinsame Marke. |",
        "| 1.2 | Beta | Zweiter Text, ganz anders. |",
        "| 1.3 | Gamma | Dritter Text. |",
    ])
    o = publish.sections_from_tables(ohne)
    assert o["1.2"].text == "Zweiter Text, ganz anders."


def test_kennung_nicht_nur_in_erster_spalte() -> None:
    """Die VDA-ISA fuehrt vor der Kennung eine Referenzspalte mit #REF!.

    Wurde nur die erste Spalte betrachtet, fiel das gesamte TISAX-Kapitel 8
    (Prototypenschutz, 23 Kriterien) auf Tabellenfuellzeichen zurueck: befuellt,
    aber ohne Inhalt — und damit schlimmer als leer, weil es wie Text aussieht.
    """
    import publish

    body = ("|  | #REF! |  | 8.1.1 |  |  | Sicherheitskonzept | "
            "Die erforderlichen Massnahmen sind umzusetzen. |")
    s = publish.sections_from_tables(body)
    assert "8.1.1" in s, list(s)
    assert "erforderlichen Massnahmen" in s["8.1.1"].text


def test_resolver_normalisiert_nur_was_er_darf() -> None:
    """Der Vergleich muss Satzform ueberstehen und Woertlaut nicht.

    Ein Resolver, der grosszuegig normalisiert, findet fast jeden Text wieder
    und belegt damit nichts. Einer, der gar nicht normalisiert, meldet
    Umbrueche als Abweichung. Beide Fehler sind hier festgenagelt.
    """
    import fundstellen as F

    # Darf angeglichen werden: Umbruch, Trennung am Zeilenende, Typografie.
    assert F.normalisiere("Informations-\nsicherheit") == "Informationssicherheit"
    assert F.normalisiere("a  b\n c") == "a b c"
    assert F.normalisiere("„Zweck“") == '"Zweck"'
    assert F.normalisiere("ﬁnden") == "finden"

    # Darf NICHT angeglichen werden: Gross-/Kleinschreibung und Verneinung
    # sind bei Normtext bedeutungstragend. MUSS und muss sind nicht dasselbe.
    assert F.normalisiere("MUSS") != F.normalisiere("muss")
    assert F.normalisiere("ist zulaessig") != F.normalisiere("ist nicht zulaessig")


def test_resolver_trennt_abweichend_von_unverifiziert() -> None:
    """Zwei Befunde, die nie zusammenfallen duerfen.

    'Unverifiziert' heisst: nicht geprueft. 'Abweichend' heisst: geprueft und
    falsch. Wer beides in ein Feld schreibt, verliert genau die Unterscheidung,
    an der der Zellversatz haengt — Kennung richtig, Wortlaut der Nachbarzeile.
    """
    import fundstellen as F

    gt = {"id": "1.1", "art": "abschnitt", "titel": "Erste Anforderung",
          "text": "Die Institution MUSS das eine tun."}

    treffer = F.pruefe_fundstelle(gt, F.normalisiere("## 1.1 Erste Anforderung\n\n"
                                                    "Die Institution MUSS das eine tun."), {}, "q")
    assert treffer.status == "verifiziert", treffer

    fehlt = F.pruefe_fundstelle(gt, F.normalisiere("## 2.1 Etwas ganz anderes"), {}, "q")
    assert fehlt.status == "unverifiziert", fehlt

    versatz = F.pruefe_fundstelle(
        gt, "", {"1.1": F.normalisiere("Erste Anforderung Die Institution SOLLTE das andere tun.")},
        "q")
    assert versatz.status == "abweichend", versatz

    # Teiltreffer ist kein Treffer: fehlt der halbe Satz, ist er nicht belegt.
    halb = F.pruefe_fundstelle(gt, F.normalisiere("1.1 Erste Anforderung Die Institution MUSS"),
                               {}, "q")
    assert halb.status == "abweichend", halb


def test_resolver_ohne_bestand_ist_nicht_gruen() -> None:
    """Ein Lauf ohne zugeordnete Bestandsdatei darf nicht wie Erfolg aussehen.

    Das ist derselbe Fehler wie die Wortdeckung, die als 100.0 startete: ein
    ungeprueftes Ergebnis, das die Form eines geprueften hat.
    """
    import fundstellen as F

    b = F.Bericht("muster", "leer")
    assert b.quote == 0.0
    assert F.passend({"kurzname": "iso27001"},
                     [Path("bsi-grundschutz.md"), Path("nis2.md")]) == []

    # Kein Raten ueber den Wortstamm: aus "bsi-kritisv" wurde einmal "bsi",
    # und die Verordnung wurde gegen "bsi-benutzerdefinierte-bausteine.md"
    # geprueft — 22 Fundstellen "unverifiziert" gegen ein fremdes Dokument.
    kandidaten = [Path("bsi-benutzerdefinierte-bausteine.md"), Path("bsi-recplast.md")]
    assert F.passend({"kurzname": "bsi-kritisv"}, kandidaten) == []

    # Die Zuordnung steht in der Ground Truth, nicht in einer Regel.
    assert F.passend({"kurzname": "bsig", "bestand_muster": ["bsi-recplast"]},
                     kandidaten) == [Path("bsi-recplast.md")]


def test_ground_truth_deckt_das_fixture() -> None:
    """Die Ground Truth und der Fixture-Generator duerfen nicht auseinanderlaufen.

    Das PDF wird aus der Ground Truth gebaut. Wuerde der Generator eigene
    Literale enthalten, pruefte der Resolver das Ergebnis gegen eine zweite,
    moeglicherweise veraltete Fassung desselben Texts — eine Selbstbestaetigung.
    """
    import json
    import re

    gt_pfad = Path(__file__).resolve().parents[1] / "fixtures" / "ground-truth" \
        / "muster-norm-99001.json"
    gt = json.loads(gt_pfad.read_text(encoding="utf-8"))
    assert len(gt["fundstellen"]) >= 10
    assert any(f["art"] == "control" for f in gt["fundstellen"])

    # Kommentare zaehlen nicht: dort darf ein Wort aus dem Primaertext stehen,
    # etwa in der Begruendung einer Spaltenbreite. Verboten ist der Woertlaut
    # als Literal, aus dem das PDF gesetzt wuerde.
    quelltext = "\n".join(
        z for z in (Path(__file__).parent / "make_fixture.py").read_text(
            encoding="utf-8").splitlines() if not z.lstrip().startswith("#"))
    for f in gt["fundstellen"]:
        for feld in ("titel", "text"):
            wert = f.get(feld, "")
            if len(wert) > 30:
                assert wert[:30] not in quelltext, \
                    f"Woertlaut steht doppelt: {feld} von {f['id']} auch im Generator"

    # Kein Wort darf so lang sein, dass der Satz es in einer Tabellenzelle
    # umbricht: genau daran zerbrach der erste Resolver-Lauf.
    for f in gt["fundstellen"]:
        if f["art"] == "control":
            laengstes = max(re.findall(r"\S+", f["titel"]), key=len)
            assert len(laengstes) <= 40, laengstes


def test_bindestrich_nur_mit_doppeltem_beleg() -> None:
    """Ein verlorener Bindestrich wird nur mit zwei Belegen zurueckgesetzt.

    Beleg 1: die Form ohne Bindestrich kommt in der Quelle nicht vor.
    Beleg 2: die Form mit Bindestrich steht dort ZUSAMMENHAENGEND.

    Ohne Beleg 2 kehrt die Reparatur die Silbentrennung um. Genau das ist
    passiert: in einem Dokument stand "Ab-\nnahme", der Extrakt hatte richtig
    "Abnahme" daraus gemacht, und die erste Fassung haette wieder "Ab-nahme"
    geschrieben — 93 Fehlalarme in einer Datei.
    """
    import re

    import verify

    quelle_roh = ("Die Verwaltung von IKT-Systemen ist geregelt.\n"
                  "Die Ab-\nnahme erfolgt spaeter. OpenLDAP bleibt OpenLDAP.")
    kompakt = re.sub(r"\s+", "", quelle_roh)

    treffer = verify.verlorene_bindestriche(
        "IKTSystemen und Abnahme und OpenLDAP", kompakt, quelle_roh)

    assert treffer == {"IKTSystemen": "IKT-Systemen"}, treffer
    assert "Abnahme" not in treffer, "Silbentrennung wurde umgekehrt"
    assert "OpenLDAP" not in treffer, "Binnenmajuskel der Quelle angetastet"


def test_unlesbares_zeichen_ist_kein_bindestrichbeleg() -> None:
    """Ein unlesbares Zeichen taugt nicht als Beleg fuer einen Bindestrich.

    Im BSIG-Druck standen 24 solche Zeichen, alle 24 waren laut amtlichem XML
    Bindestriche — und daraus wurde kurzzeitig eine Regel. Sie hielt genau ein
    Dokument weit: in einer anderen Datei desselben Bestandes stehen 1639
    davon, und dort sind es Trennstriche am Zeilenende. Dasselbe Zeichen, zwei
    Bedeutungen. Gemeldet statt geraten.
    """
    import verify

    roh = "Die Ab\ufffenahme des IKT\ufffeSystems."
    assert verify.unlesbar_im_wort(roh) == 2
    # Die Belegquellen lassen das Zeichen stehen, machen also keinen
    # Bindestrich daraus.
    assert "-" not in verify.zusammenhaengende_quelle("", roh)
    assert "-" not in verify.quelle_kompakt("", roh)


def test_front_matter_ueberlebt_die_reparatur() -> None:
    """Der Kopf muss nach dem Vermerk noch an erster Stelle stehen.

    Diese Zeile hat gefehlt, und ein einziger vertauschter Variablenname
    ("roh" hielt ploetzlich den PDF-Text statt des Markdowns) hat den Kopf von
    187 Extrakten aus dem falschen Text geschnitten. Der Schaden faellt beim
    Schreiben nicht auf — nur beim naechsten Lesen.
    """
    import bindestriche
    import publish

    roh = ('---\nsource_file: "X.pdf"\npages: 3\nextraction_status: ok\n---\n'
           "\n<!-- ACSOS -->\n\nText mit IKT-Systemen.\n")
    meta, koerper = publish.split_front_matter(roh)
    kopf = roh[:len(roh) - len(koerper)]
    ergebnis = bindestriche.vermerke(kopf, 2, "a -> b") + koerper

    assert ergebnis.startswith("---\n"), ergebnis[:60]
    neu_meta, neu_koerper = publish.split_front_matter(ergebnis)
    assert neu_meta["source_file"] == "X.pdf"
    assert neu_meta["pages"] == "3"
    assert neu_meta["restored_hyphens"] == "2"
    assert neu_koerper == koerper, "Rumpf veraendert"


def test_resolver_erlaubt_einschuebe_aber_keine_luecke() -> None:
    """Eingeschobenes Fremdmaterial ist erlaubt, fehlender Woertlaut nicht.

    Ein Gesetzesdruck schiebt mitten in den Absatz Seitenkopf, Fusszeile und
    Verweise ein. Wer am Stueck vergleicht, meldet jeden laengeren Paragrafen
    als Abweichung — und nach dem dritten Fehlalarm liest niemand mehr hin.
    """
    import fundstellen as F

    soll = "Die Institution MUSS die Anlage schuetzen und den Vorfall melden."
    mit_einschub = ("Die Institution MUSS die Anlage schuetzen "
                    "Seite 4 von 12 Nichtamtliches Inhaltsverzeichnis "
                    "und den Vorfall melden.")
    ok, einschuebe, fehlt = F.enthalten_mit_einschueben(soll, mit_einschub)
    assert ok and einschuebe >= 1 and not fehlt

    ohne_wort = "Die Institution MUSS die Anlage schuetzen und den melden."
    ok2, _, fehlt2 = F.enthalten_mit_einschueben(soll, ohne_wort)
    assert not ok2 and "Vorfall" in fehlt2


def test_resolver_nimmt_nicht_das_inhaltsverzeichnis() -> None:
    """Der erste Ankertreffer ist im Gesetzesdruck regelmaessig der falsche.

    Vorn steht ein Inhaltsverzeichnis mit denselben Ueberschriften. Wer dort
    stehenbleibt, meldet Woerter als fehlend, die zwanzig Seiten weiter unten
    stehen — so entstanden 13 Fehlalarme.
    """
    import fundstellen as F

    soll = "Die Institution MUSS die kritische Anlage besonders schuetzen."
    bestand = ("Inhaltsuebersicht Die Institution MUSS die kritische Anlage ... 12 "
               + "Fuelltext " * 60
               + "Die Institution MUSS die kritische Anlage besonders schuetzen.")
    ok, _, fehlt = F.enthalten_mit_einschueben(soll, bestand)
    assert ok, fehlt


def test_rechtsakt_klebt_die_aufzaehlungsmarke_nicht_an() -> None:
    """Aus <DT>1.</DT><DD>Konzepte</DD> muss "1. Konzepte" werden.

    Ohne Leerzeichen an der Elementgrenze entstand "1.Konzepte" — und der
    Resolver meldete jeden Absatz mit Aufzaehlung als Abweichung, obwohl der
    Extrakt richtig war. Ein Pruefmassstab, der selbst falsch zusammensetzt,
    erzeugt genau die Fehlalarme, die ihn unbrauchbar machen.
    """
    import xml.etree.ElementTree as ET

    import rechtsakte

    el = ET.fromstring("<Content><P>Folgendes umfassen: <DL><DT>1.</DT>"
                       "<DD>Konzepte zur Risikoanalyse,</DD><DT>2.</DT>"
                       "<DD>Bewaeltigung von Vorfaellen.</DD></DL></P></Content>")
    text = rechtsakte.blocktext(el)
    assert "1. Konzepte" in text, text
    assert "2. Bewaeltigung" in text, text
    assert "1.Konzepte" not in text


def test_alle_ground_truths_sind_lesbar() -> None:
    """Jede Ground-Truth-Datei muss gueltig und vollstaendig deklariert sein.

    Ein Primaertext, der sich nicht laden laesst oder seine Zuordnung nicht
    nennt, faellt sonst erst im Bericht auf — als "unverifiziert", also als
    Aussage ueber die Daten statt ueber das Werkzeug.
    """
    import json

    ordner = Path(__file__).resolve().parents[1] / "fixtures" / "ground-truth"
    dateien = sorted(ordner.glob("*.json"))
    assert dateien, "kein Primaertext vorhanden"
    for datei in dateien:
        gt = json.loads(datei.read_text(encoding="utf-8"))
        for feld in ("quelle", "kurzname", "herkunft", "fundstellen"):
            assert feld in gt, f"{datei.name}: {feld} fehlt"
        assert isinstance(gt.get("bestand_muster", []), list), datei.name
        assert gt["fundstellen"], f"{datei.name}: keine Fundstellen"
        for f in gt["fundstellen"]:
            assert f.get("id"), f"{datei.name}: Fundstelle ohne Kennung"
            assert f.get("text") or f.get("titel"), f"{datei.name}: {f['id']} ohne Inhalt"


def test_vollstaendigkeit_findet_die_nie_eingelesene_quelle(tmp_path: Path) -> None:
    """Der Waechter muss melden, was FEHLT — nicht pruefen, was da ist.

    Alle anderen Pruefungen sehen sich Extrakte an: Deckung, Zellversatz,
    Kennungen, Fundstellen. Ein Dokument, das nie eingespeist wurde, hat keinen
    Extrakt, der auffallen koennte, und keinen Eintrag, der widerspraeche — es
    faellt durch jedes Netz. So fehlten 134 von 602 Dokumenten, und aufgefallen
    ist es erst ausserhalb dieses Systems.

    Beide Richtungen: bei vollstaendigem Bestand schweigen, bei einer fehlenden
    Quelle anschlagen — auch wenn sie in einem Archiv steckt.
    """
    import json
    import zipfile
    from subprocess import run

    wurzel = Path(__file__).resolve().parents[1]
    eingang = tmp_path / "input"
    eingang.mkdir()
    (eingang / "vorhanden.pdf").write_bytes(b"%PDF-1.4 nur ein Platzhalter")
    korpus = tmp_path / "_KORPUS.json"

    def lauf() -> tuple[int, str]:
        e = run([sys.executable, str(wurzel / "vollstaendigkeit.py"),
                 "--input", str(eingang), "--korpus", str(korpus), "--strict"],
                capture_output=True, text=True, cwd=wurzel)
        return e.returncode, e.stdout

    korpus.write_text(json.dumps(
        {"documents": [{"slug": "vorhanden", "source_file": "vorhanden.pdf"}]}), encoding="utf-8")
    code, aus = lauf()
    assert code == 0, f"schlug bei vollstaendigem Bestand an:\n{aus}"

    # Eine zweite Quelle, die niemand konvertiert hat.
    (eingang / "vergessen.pdf").write_bytes(b"%PDF-1.4 auch ein Platzhalter")
    code, aus = lauf()
    assert code == 1, f"die fehlende Quelle blieb unbemerkt:\n{aus}"
    assert "vergessen.pdf" in aus

    # Und eine, die in einem Archiv steckt: 130 Dokumente lagen so daneben.
    (eingang / "vergessen.pdf").unlink()
    with zipfile.ZipFile(eingang / "paket.zip", "w") as z:
        z.writestr("Checkliste_APP.1.1.xlsx", "x")
    code, aus = lauf()
    assert code == 1, f"das nicht ausgepackte Archiv blieb unbemerkt:\n{aus}"
    assert "Checkliste_APP.1.1.xlsx" in aus

    # Ein byte-identisches Duplikat unter anderem Namen (ZIP-Inhalt neben dem
    # Original) ist keine Luecke: der Hash belegt, dass der Inhalt im Bestand ist.
    # Ein gleichnamiges Duplikat mit ANDEREM Inhalt bleibt aber eine Luecke.
    (eingang / "paket.zip").unlink()
    import hashlib
    inhalt = (eingang / "vorhanden.pdf").read_bytes()
    korpus.write_text(json.dumps({"documents": [{
        "slug": "vorhanden", "source_file": "vorhanden.pdf",
        "source_sha256": hashlib.sha256(inhalt).hexdigest()}]}), encoding="utf-8")
    (eingang / "kopie").mkdir()
    (eingang / "kopie" / "vorhanden-2023.pdf").write_bytes(inhalt)
    code, aus = lauf()
    assert code == 0, f"das byte-identische Duplikat wurde als Luecke gemeldet:\n{aus}"
    assert "Duplikat" in aus
    (eingang / "kopie" / "vorhanden-2023.pdf").write_bytes(inhalt + b" geaendert")
    code, aus = lauf()
    assert code == 1, f"die abweichende Fassung blieb unbemerkt:\n{aus}"


def test_seitenmarken_muessen_lueckenlos_sein(tmp_path: Path) -> None:
    """Jede Seite eines PDF-Extrakts braucht ihre Marke.

    Laenge und Deckung sehen eine fehlende Seitenmarke nicht: die Woerter sind
    da, nur nicht dort, wo ein Zitat mit Seitenzahl sie sucht. Beide
    Richtungen: lueckenlos schweigt, eine Luecke schlaegt an.
    """
    import json

    import pruefe

    md = tmp_path / "doc.md"
    md.write_text("---\npages: 3\n---\n<!-- page: 1 -->\nA\n<!-- page: 2 -->\nB\n"
                  "<!-- page: 3 -->\nC\n", encoding="utf-8")
    reg = tmp_path / "_KORPUS.json"
    eintrag = {"slug": "doc", "source_file": "doc.pdf", "markdown": str(md),
               "pages": 3, "words": 300, "text_coverage_percent": 100.0, "woertlich": True}
    reg.write_text(json.dumps({"documents": [eintrag]}), encoding="utf-8")

    b = pruefe.Bericht()
    pruefe.pruefe_korpus(reg, b)
    assert not [f for f in b.befunde if "Marke" in f.aussage], b.befunde

    md.write_text("---\npages: 3\n---\n<!-- page: 1 -->\nA\n<!-- page: 3 -->\nC\n",
                  encoding="utf-8")
    b = pruefe.Bericht()
    pruefe.pruefe_korpus(reg, b)
    treffer = [f for f in b.befunde if "Marke" in f.aussage]
    assert treffer and "[2]" in treffer[0].zahl, b.befunde


# Muss am Dateiende stehen. Stand dieser Block frueher in der Mitte, war die
# Datei beim Aufruf von main() nur bis dorthin ausgefuehrt: alles danach
# definierte existierte noch nicht und lief im Skriptpfad nie mit.
if __name__ == "__main__":
    raise SystemExit(main())
